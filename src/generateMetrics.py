"""
Author: Bhargav D V, Research Scholar, IIITB, under guidance of Prof. Madhav Rao.
This script is used to generate Metrics
"""


#------------import modules----------------#
from optimizationUtils import determineDecisionVariableLimit
from pymoo.core.problem import Problem
from globalVariables import *
import numpy as np
import re
import os
#from circuits.UnsignedCarryLookaheadAdder8bit import Circuit
import importlib
from hardwareMetricsEstimation import runSynthesisFinal
import math
from openpyxl import Workbook
#------------import modules----------------#



def countFiles(directory_path):
    count = 0
    for entry in os.listdir(directory_path):
        full_path = os.path.join(directory_path, entry)
        if os.path.isfile(full_path):
            count += 1
    return count



def updateTCLCommands(yosysCmd,openstaCmd,Z):
        
        YOSYS_SCRIPT_EVOLUTION="""
#---------------- yosys commands -------------------
#import yosys environment
yosys -import

#read liberty file for required technology
read_liberty -lib {}

#read verilog design file
{}
read_verilog -overwrite {}{}_{}.v

hierarchy -check -top {}_{}

#store the area result in a file
tee -q -o {}{}_result_{}.txt stat -liberty {}

#exit
exit


#---------------- yosys commands -------------------
""".format(LIBERTY_FILE,yosysCmd,CIRCUIT,DESIGN,str(Z),DESIGN,str(Z),EVOLUTION_PATH,DESIGN,str(Z),LIBERTY_FILE)
        
        OPENSTA_SCRIPT_EVOLUTION="""
#----------------- opensta commands ----------------------
#----------estimating power in opensta-----------
#import the library file
read_liberty {}

#read the design file
{}
read_verilog {}{}_{}.v 


#identify the top module name
link_design {}_{}

#clock frequency
create_clock -period  {} -name clk

#output load
set_load -pin_load 2.479712 [lsearch -inline -all -not -exact [all_outputs] clk]

set_driving_cell -lib_cell INVx1_ASAP7_75t_R [lsearch -inline -all -not -exact [all_inputs] clk]

#switching factor for truly random data
set_power_activity -input_ports [lsearch -inline -all -not -exact [all_inputs] clk] -activity 0.25


#input output delay constraint
set_input_delay 0 -clock clk [lsearch -inline -all -not -exact [all_inputs] clk]
set_output_delay 0 -clock clk [lsearch -inline -all -not -exact [all_outputs] clk]

#estimate power
report_power >> {}{}_result_{}.txt

#----------estimating power in opensta-----------


#----------------estimating critical path delay -------------------


#critical path delay
report_checks -unconstrained >> {}{}_result_{}.txt
#----------------estimating critical path delay -------------------

#exit
exit

#----------------- opensta commands ----------------------
""".format(LIBERTY_FILE,openstaCmd,CIRCUIT,DESIGN,str(Z),DESIGN,str(Z),str(CLOCK_PERIOD),EVOLUTION_PATH,DESIGN,str(Z),EVOLUTION_PATH,DESIGN,str(Z))
        
        return YOSYS_SCRIPT_EVOLUTION,OPENSTA_SCRIPT_EVOLUTION


def runFramework():
   

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1,column=1,value='Metrics {}'.format(DESIGN))
    ws.cell(row=2,column=1,value='DESIGN')
    ws.cell(row=2,column=2,value='AREA(um^2)')
    ws.cell(row=2,column=3,value='DELAY(ps)')
    ws.cell(row=2,column=4,value='INTERNAL POWER(W)')
    ws.cell(row=2,column=5,value='SWITCHING POWER(W)')
    ws.cell(row=2,column=6,value='LEAKAGE POWER(W)')
    ws.cell(row=2,column=7,value='TOTAL POWER(W)')
    
    
    DESIGNS=countFiles(CIRCUIT)

    
    for unit in range(DESIGNS):
        print('=================== {} {} ======================'.format(DESIGN,str(unit)))
        #input('press')
        

        verilogFile=open(CIRCUIT+DESIGN+f'_{unit}.v',mode='r',encoding='utf-8')
        verilogCode=verilogFile.read()
        verilogFile.close()
        


        # Regular expressions for both patterns
        nr_pattern = re.findall(r'NR_\d+_\d+', verilogCode)
        custom_adder_pattern = re.findall(r'customAdder\d+_\d+', verilogCode)

        yosysCmd = ''                                                                                                  
        #add path of modules here for non recursive multiplier
        for file_name in set(nr_pattern):
            yosysCmd += f"read_verilog -overwrite {MODULES_PATH}{file_name}.v\n"
        
        
        #add path of modules here for adder
        for file_name in set(custom_adder_pattern):
            
            yosysCmd += f"read_verilog -overwrite {MODULES_PATH}{file_name}.v\n"

        

        openstaCmd = ''                                                                                                  
        #add path of modules here for non recursive multiplier
        for file_name in set(nr_pattern):
            openstaCmd += f"read_verilog {MODULES_PATH}{file_name}.v\n"
        
        
        #add path of modules here for adder
        for file_name in set(custom_adder_pattern):
            
            openstaCmd += f"read_verilog {MODULES_PATH}{file_name}.v\n"
            
        

        YOSYS_SCRIPT_EVOLUTION,OPENSTA_SCRIPT_EVOLUTION=updateTCLCommands(yosysCmd,openstaCmd,unit)

        totalArea,totalDelay,interalPower,switchingPower,leakagePower,totalPower=runSynthesisFinal(YOSYS_SCRIPT_EVOLUTION,OPENSTA_SCRIPT_EVOLUTION,unit)



        ws.cell(row=unit+3,column=1,value='{}_{}'.format(str(DESIGN),str(unit)))
        ws.cell(row=unit+3,column=2,value=float(totalArea))
        ws.cell(row=unit+3,column=3,value=float(totalDelay))
        ws.cell(row=unit+3,column=4,value=float(interalPower))
        ws.cell(row=unit+3,column=5,value=float(switchingPower))
        ws.cell(row=unit+3,column=6,value=float(leakagePower))
        ws.cell(row=unit+3,column=7,value=float(totalPower))
        

    wb.save('{}metrics.xlsx'.format(METRICS_PATH))
    print('{}metrics.xlsx'.format(METRICS_PATH))
# /home/niranjan_yash/Emergin_Mult/Project_Latest/project/src/generateMetrics.py
runFramework()